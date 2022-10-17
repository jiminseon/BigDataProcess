#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

grade = dict()
row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		total = ws.cell(row = row_id, column = 7).value
		grade[row_id] = total
	row_id += 1

grade2 = sorted(grade.items(), key = lambda x : x[1], reverse = True)
dic = dict(grade2)
lst = list(dic.keys())
print(lst)

row_id = i = 0
num = ws.max_row - 1
a = int(num * 0.3 // 2)
b = int((num * 0.7 - a) // 2)
c = (num - (a + b) * 2) // 2 
a1 = b1 = c1 = 0 
a2 = b2 = c2 = -1
same = 0

for row in ws:
	if row_id != 1:
		if a1 < a:
			if lst[i] == lst[i+1]:
				same += 1
			else:
				for (i - same) in same:
					ws.cell(row = int(lst[i]), column = 8).value = 'A+'
				ws.cell(row = int(lst[i]), column = 8).value = 'A+'
				same = 0
			a1 += 1
			i += 1
			row_id += 1
			continue
		elif a2 < a:
			if lst[i] == lst[i+1]:
				same += 1
			else:
				for (i - same) in same:
					ws.cell(row = int(lst[i]), column = 8).value = 'A0'
				ws.cell(row = int(lst[i]), column = 8).value = 'A0'
				same = 0
			a2 += 1
			i += 1
			row_id += 1
			continue
		elif b1 < b:
			if lst[i] == lst[i+1]:
				same += 1
			else:
				for (i - same) in same:
					ws.cell(row = int(lst[i]), column = 8).value = 'B+'
				ws.cell(row = int(lst[i]), column = 8).value = 'B+'
				same = 0
			ws.cell(row = int(lst[i]), column = 8).value = 'B+'
			b1 += 1
			i += 1
			row_id += 1
			continue
		elif b2 < b:
			ws.cell(row = int(lst[i]), column = 8).value = 'B0'
			b2 += 1
			i += 1
			row_id += 1
			continue
		elif c1 < c:
			ws.cell(row = int(lst[i]), column = 8).value = 'C+'
			c1 += 1
			i += 1
			row_id += 1
			continue
		elif c2 < c:
			ws.cell(row = int(lst[i]), column = 8).value = 'C0'
			c2 += 1
			i += 1
			row_id += 1
			continue
	row_id += 1
wb.save("student.xlsx")
	

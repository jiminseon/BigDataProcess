#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

grade = dict()
row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		total = ws.cell(row = row_id, column = 7).value
		grade[row_id] = total
	row_id += 1

grade2 = sorted(grade.items(), key = lambda x : x[1], reverse = True)
dic = dict(grade2)
lst = list(dic.keys())

row_id = i = 0
num = ws.max_row - 1
a = int(num * 0.3) 
b = int(num * 0.4) 
c = num - (a + b)

a1 = a2 = b1 = b2 = c1 = c2 = 0
if (a % 2 != 0):
	a1 = 0
	a2 = -1
if (b % 2 != 0):
	b1 = 0
	b2 = -1
if (c % 2 != 0):
	c1 = 0
	c2 = -1
a = a // 2
b = b // 2
c = c // 2
print(a, b, c)
for row in ws:
	if row_id != 1:
		if a1 < a:
			ws.cell(row = int(lst[i]), column = 8).value = 'A+'
			a1 += 1
			i += 1
			row_id += 1
			continue
		elif a2 < a:
			ws.cell(row = int(lst[i]), column = 8).value = 'A0'
			a2 += 1
			i += 1
			row_id += 1
			continue
		elif b1 < b:
			ws.cell(row = int(lst[i]), column = 8).value = 'B+'
			b1 += 1
			i += 1
			row_id += 1
			continue
		elif b2 < b:
			ws.cell(row = int(lst[i]), column = 8).value = 'B0'
			b2 += 1
			i += 1
			row_id += 1
			continue
		elif c1 < c:
			ws.cell(row = int(lst[i]), column = 8).value = 'C+'
			c1 += 1
			i += 1
			row_id += 1
			continue
		elif c2 < c:
			ws.cell(row = int(lst[i]), column = 8).value = 'C0'
			c2 += 1
			i += 1
			row_id += 1
			continue
	row_id += 1
wb.save("student.xlsx")
	
